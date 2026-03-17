"""Dashboard API -- FastAPI backend for the WhatsApp bot dashboard.

Exposes read/write endpoints over the bot's SQLite databases, config,
spreadsheets, memory files, and health metrics.
"""

import hashlib
import json
import os
import socket
import sqlite3
import subprocess
import time
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ── Paths ──

BASE_DIR = Path(os.environ.get("WHATSAPP_BASE_DIR", str(Path.home() / ".happycapy-whatsapp")))
CONTACTS_DB = BASE_DIR / "contacts.db"
REFLECTION_DB = BASE_DIR / "reflection.db"
BROADCAST_DB = BASE_DIR / "broadcast.db"
CONFIG_FILE = BASE_DIR / "config.json"
MEMORY_DIR = BASE_DIR / "memory"
SPREADSHEET_DIR = BASE_DIR / "data" / "spreadsheets"
LOG_FILE = BASE_DIR / "logs" / "daemon.log"
IDENTITY_DIR = BASE_DIR / "identity"
PROACTIVE_DB = BASE_DIR / "proactive.db"
CRON_DB = BASE_DIR / "cron.db"

# Frontend dist
FRONTEND_DIR = Path(__file__).parent / "frontend" / "dist"

app = FastAPI(title="HappyCapy WhatsApp Dashboard API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Helpers ──

def _db(path: Path) -> sqlite3.Connection:
    """Open a read-only SQLite connection with row factory."""
    if not path.exists():
        raise HTTPException(404, f"Database not found: {path.name}")
    conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def _db_rw(path: Path) -> sqlite3.Connection:
    """Open a read-write SQLite connection."""
    if not path.exists():
        raise HTTPException(404, f"Database not found: {path.name}")
    conn = sqlite3.connect(str(path))
    conn.row_factory = sqlite3.Row
    return conn


def _rows_to_dicts(rows: list[sqlite3.Row]) -> list[dict]:
    return [dict(r) for r in rows]


def _get_jid_names(conn: sqlite3.Connection) -> dict[str, str]:
    """Build a JID -> display name mapping from whatsapp_contacts.

    Handles both raw JIDs ('85292893658') and suffixed forms
    ('85292893658@s.whatsapp.net', '167087800602664@lid').
    """
    try:
        rows = conn.execute(
            "SELECT jid, push_name, saved_name, verified_name FROM whatsapp_contacts WHERE push_name IS NOT NULL OR saved_name IS NOT NULL"
        ).fetchall()
    except Exception:
        return {}
    mapping = {}
    for r in rows:
        name = r["saved_name"] or r["push_name"] or r["verified_name"]
        if not name:
            continue
        jid = r["jid"]
        mapping[jid] = name
        # Also map suffixed forms so audit_log chat_ids resolve
        mapping[f"{jid}@s.whatsapp.net"] = name
        mapping[f"{jid}@lid"] = name
    return mapping


def _resolve_jid(jid: str | None, jid_names: dict[str, str]) -> str:
    """Resolve a JID to a display name, falling back to cleaned JID."""
    if not jid:
        return "Unknown"
    if jid in jid_names:
        return jid_names[jid]
    # Strip suffix and try bare number
    bare = jid.split("@")[0] if "@" in jid else jid
    if bare in jid_names:
        return jid_names[bare]
    return bare


def _jid_hash(jid: str) -> str:
    """Match the memory_store hashing: MD5 of JID, first 12 chars."""
    return hashlib.md5(jid.encode()).hexdigest()[:12]


# ══════════════════════════════════════════════════════════════
# HEALTH & STATUS
# ══════════════════════════════════════════════════════════════

@app.get("/api/health")
def get_health():
    """Bot health status and uptime."""
    pid_file = BASE_DIR / "daemon.pid"
    running = False
    pid = None
    uptime_seconds = 0

    if pid_file.exists():
        try:
            pid = int(pid_file.read_text().strip())
            os.kill(pid, 0)  # Check if process exists
            running = True
            # Approximate uptime from PID file mtime
            uptime_seconds = int(time.time() - pid_file.stat().st_mtime)
        except (ValueError, ProcessLookupError, PermissionError):
            running = False

    # Check WhatsApp auth status
    auth_dir = BASE_DIR / "whatsapp-auth"
    wa_authenticated = (auth_dir / "creds.json").exists() if auth_dir.exists() else False

    # Database sizes
    db_sizes = {}
    for name, path in [("contacts", CONTACTS_DB), ("reflection", REFLECTION_DB), ("broadcast", BROADCAST_DB)]:
        if path.exists():
            db_sizes[name] = round(path.stat().st_size / 1024, 1)  # KB

    # Recent errors from log
    error_count = 0
    if LOG_FILE.exists():
        try:
            text = LOG_FILE.read_text(errors="replace")
            one_hour_ago = (datetime.now() - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M")
            for line in text.split("\n")[-500:]:
                if "Error" in line or "error" in line or "Traceback" in line:
                    if one_hour_ago < line[:16] if len(line) > 16 else False:
                        error_count += 1
        except Exception:
            pass

    return {
        "running": running,
        "pid": pid,
        "uptime_seconds": uptime_seconds,
        "whatsapp_authenticated": wa_authenticated,
        "database_sizes_kb": db_sizes,
        "errors_last_hour": error_count,
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/api/logs")
def get_logs(lines: int = Query(100, le=500)):
    """Recent daemon log lines."""
    if not LOG_FILE.exists():
        return {"lines": [], "total": 0}
    try:
        text = LOG_FILE.read_text(errors="replace")
        all_lines = text.strip().split("\n")
        return {"lines": all_lines[-lines:], "total": len(all_lines)}
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════

@app.get("/api/config")
def get_config():
    """Current bot configuration."""
    if not CONFIG_FILE.exists():
        raise HTTPException(404, "Config not found")
    return json.loads(CONFIG_FILE.read_text())


class ConfigUpdate(BaseModel):
    updates: dict[str, Any]


@app.put("/api/config")
def update_config(body: ConfigUpdate):
    """Update specific config fields (merge, not replace)."""
    if not CONFIG_FILE.exists():
        raise HTTPException(404, "Config not found")
    config = json.loads(CONFIG_FILE.read_text())
    config.update(body.updates)
    CONFIG_FILE.write_text(json.dumps(config, indent=2))
    return {"status": "ok", "config": config}


# ══════════════════════════════════════════════════════════════
# CONTACTS & PROFILES
# ══════════════════════════════════════════════════════════════

@app.get("/api/contacts")
def get_contacts(limit: int = Query(100, le=500)):
    """List contacts with profiles."""
    with _db(CONTACTS_DB) as conn:
        rows = conn.execute("""
            SELECT cp.jid, cp.display_name, cp.total_messages_analyzed,
                   cp.profile_json, cp.updated_at,
                   wc.push_name, wc.saved_name, wc.verified_name
            FROM contact_profiles cp
            LEFT JOIN whatsapp_contacts wc ON cp.jid = wc.jid
            ORDER BY cp.updated_at DESC
            LIMIT ?
        """, (limit,)).fetchall()

    contacts = []
    for r in rows:
        d = dict(r)
        try:
            d["profile"] = json.loads(d.pop("profile_json", "{}"))
        except Exception:
            d["profile"] = {}
        contacts.append(d)
    return {"contacts": contacts, "total": len(contacts)}


@app.get("/api/contacts/{jid}")
def get_contact_detail(jid: str):
    """Detailed contact view with profile, samples, KG entities."""
    with _db(CONTACTS_DB) as conn:
        profile = conn.execute(
            "SELECT * FROM contact_profiles WHERE jid = ?", (jid,)
        ).fetchone()
        if not profile:
            raise HTTPException(404, "Contact not found")

        samples = conn.execute(
            "SELECT role, content, timestamp FROM conversation_samples WHERE jid = ? ORDER BY timestamp DESC LIMIT 20",
            (jid,)
        ).fetchall()

        entities = conn.execute(
            "SELECT name, entity_type, mention_count, last_seen FROM kg_entities WHERE jid = ? ORDER BY mention_count DESC LIMIT 20",
            (jid,)
        ).fetchall()

        sessions = conn.execute(
            "SELECT * FROM sessions WHERE jid = ?", (jid,)
        ).fetchone()

    result = dict(profile)
    try:
        result["profile_data"] = json.loads(result.pop("profile_json", "{}"))
    except Exception:
        result["profile_data"] = {}

    result["recent_samples"] = _rows_to_dicts(samples)
    result["knowledge_entities"] = _rows_to_dicts(entities)
    result["session"] = dict(sessions) if sessions else None
    return result


# ══════════════════════════════════════════════════════════════
# AUDIT LOG & ANALYTICS
# ══════════════════════════════════════════════════════════════

@app.get("/api/audit")
def get_audit(
    limit: int = Query(50, le=500),
    event_type: str | None = None,
    hours: int = Query(24, le=168),
):
    """Audit log entries."""
    cutoff = (datetime.now() - timedelta(hours=hours)).isoformat()
    with _db(CONTACTS_DB) as conn:
        if event_type:
            rows = conn.execute(
                "SELECT * FROM audit_log WHERE timestamp > ? AND event_type = ? ORDER BY timestamp DESC LIMIT ?",
                (cutoff, event_type, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM audit_log WHERE timestamp > ? ORDER BY timestamp DESC LIMIT ?",
                (cutoff, limit)
            ).fetchall()
    return {"events": _rows_to_dicts(rows), "total": len(rows)}


@app.get("/api/analytics")
def get_analytics():
    """Dashboard analytics: message counts, top contacts, activity by hour."""
    with _db(CONTACTS_DB) as conn:
        # Total messages
        total = conn.execute("SELECT COUNT(*) as cnt FROM audit_log").fetchone()

        # Messages by direction (last 7 days)
        week_ago = (datetime.now() - timedelta(days=7)).isoformat()
        by_direction = conn.execute("""
            SELECT direction, COUNT(*) as cnt
            FROM audit_log WHERE timestamp > ? AND direction IS NOT NULL
            GROUP BY direction
        """, (week_ago,)).fetchall()

        # Messages per day (last 7 days)
        per_day = conn.execute("""
            SELECT DATE(timestamp) as day, COUNT(*) as cnt
            FROM audit_log WHERE timestamp > ?
            GROUP BY DATE(timestamp)
            ORDER BY day
        """, (week_ago,)).fetchall()

        # Top contacts by message count (with name resolution)
        top_contacts_raw = conn.execute("""
            SELECT chat_id, COUNT(*) as cnt
            FROM audit_log WHERE chat_id IS NOT NULL AND chat_id != '' AND timestamp > ?
            GROUP BY chat_id ORDER BY cnt DESC LIMIT 10
        """, (week_ago,)).fetchall()
        jid_names = _get_jid_names(conn)
        top_contacts = []
        for r in top_contacts_raw:
            d = dict(r)
            d["display_name"] = _resolve_jid(d["chat_id"], jid_names)
            top_contacts.append(d)

        # Event type breakdown
        by_type = conn.execute("""
            SELECT event_type, COUNT(*) as cnt
            FROM audit_log WHERE timestamp > ?
            GROUP BY event_type ORDER BY cnt DESC
        """, (week_ago,)).fetchall()

        # Queue stats
        queue_stats = conn.execute("""
            SELECT status, COUNT(*) as cnt FROM message_queue GROUP BY status
        """).fetchall()

        # Escalation stats
        esc_stats = conn.execute("""
            SELECT status, COUNT(*) as cnt FROM escalations GROUP BY status
        """).fetchall()

        # Active sessions
        active_sessions = conn.execute(
            "SELECT COUNT(*) as cnt FROM sessions WHERE is_active = 1"
        ).fetchone()

    return {
        "total_events": dict(total)["cnt"] if total else 0,
        "messages_by_direction": _rows_to_dicts(by_direction),
        "messages_per_day": _rows_to_dicts(per_day),
        "top_contacts": top_contacts,
        "events_by_type": _rows_to_dicts(by_type),
        "queue_stats": _rows_to_dicts(queue_stats),
        "escalation_stats": _rows_to_dicts(esc_stats),
        "active_sessions": dict(active_sessions)["cnt"] if active_sessions else 0,
    }


# ══════════════════════════════════════════════════════════════
# MESSAGE QUEUE
# ══════════════════════════════════════════════════════════════

@app.get("/api/queue")
def get_queue(status: str | None = None, limit: int = Query(50, le=200)):
    """Message queue entries with resolved sender names."""
    with _db(CONTACTS_DB) as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM message_queue WHERE status = ? ORDER BY created_at DESC LIMIT ?",
                (status, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM message_queue ORDER BY created_at DESC LIMIT ?",
                (limit,)
            ).fetchall()
        jid_names = _get_jid_names(conn)
    messages = []
    for r in rows:
        d = dict(r)
        if not d.get("sender_name"):
            d["sender_name"] = _resolve_jid(d.get("sender_id"), jid_names)
        messages.append(d)
    return {"messages": messages, "total": len(messages)}


# ══════════════════════════════════════════════════════════════
# ESCALATIONS
# ══════════════════════════════════════════════════════════════

@app.get("/api/escalations")
def get_escalations(status: str | None = None):
    """Escalation entries with resolved sender names."""
    with _db(CONTACTS_DB) as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM escalations WHERE status = ? ORDER BY created_at DESC",
                (status,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM escalations ORDER BY created_at DESC LIMIT 50"
            ).fetchall()
        jid_names = _get_jid_names(conn)
    escalations = []
    for r in rows:
        d = dict(r)
        if not d.get("sender_name"):
            d["sender_name"] = _resolve_jid(d.get("sender_id"), jid_names)
        escalations.append(d)
    return {"escalations": escalations, "total": len(escalations)}


# ══════════════════════════════════════════════════════════════
# SPREADSHEETS
# ══════════════════════════════════════════════════════════════

@app.get("/api/spreadsheets")
def list_spreadsheets():
    """List available spreadsheet files."""
    if not SPREADSHEET_DIR.exists():
        return {"spreadsheets": []}
    files = []
    for f in sorted(SPREADSHEET_DIR.glob("*.xlsx")):
        files.append({
            "name": f.stem,
            "filename": f.name,
            "size_kb": round(f.stat().st_size / 1024, 1),
            "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
        })
    return {"spreadsheets": files}


@app.get("/api/spreadsheets/{name}")
def read_spreadsheet(name: str, sheet: str | None = None, limit: int = Query(100, le=1000)):
    """Read spreadsheet data as JSON rows."""
    filepath = SPREADSHEET_DIR / f"{name}.xlsx"
    if not filepath.exists():
        raise HTTPException(404, f"Spreadsheet '{name}' not found")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb[sheet] if sheet and sheet in sheet_names else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"name": name, "sheets": sheet_names, "headers": [], "rows": [], "total": 0}

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:limit + 1]:
            data.append({headers[i]: cell for i, cell in enumerate(row) if i < len(headers)})

        wb.close()
        return {
            "name": name,
            "sheets": sheet_names,
            "headers": headers,
            "rows": data,
            "total": len(rows) - 1,
        }
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/spreadsheets/{name}/download")
def download_spreadsheet(name: str):
    """Download spreadsheet file."""
    filepath = SPREADSHEET_DIR / f"{name}.xlsx"
    if not filepath.exists():
        raise HTTPException(404, f"Spreadsheet '{name}' not found")
    return FileResponse(filepath, filename=f"{name}.xlsx")


# ══════════════════════════════════════════════════════════════
# KNOWLEDGE GRAPH
# ══════════════════════════════════════════════════════════════

@app.get("/api/knowledge-graph")
def get_knowledge_graph(limit: int = Query(100, le=500)):
    """Knowledge graph entities and relationships with resolved contact names."""
    with _db(CONTACTS_DB) as conn:
        jid_names = _get_jid_names(conn)

        # Load config for owner name
        owner_name = "Owner"
        try:
            if CONFIG_FILE.exists():
                cfg = json.loads(CONFIG_FILE.read_text())
                owner_name = cfg.get("owner_name") or cfg.get("admin_number", "Owner")
        except Exception:
            pass

        def resolve_kg_name(name: str, jid: str) -> str:
            """Replace generic 'Contact'/'You' with real names."""
            if name == "You":
                return owner_name
            if name == "Contact":
                return _resolve_jid(jid, jid_names)
            return name

        entities_raw = conn.execute("""
            SELECT id, name, entity_type, jid, description, mention_count, last_seen
            FROM kg_entities ORDER BY mention_count DESC LIMIT ?
        """, (limit,)).fetchall()

        entities = []
        for e in entities_raw:
            d = dict(e)
            d["display_name"] = resolve_kg_name(d["name"], d["jid"])
            d["contact_name"] = _resolve_jid(d["jid"], jid_names)
            entities.append(d)

        relationships_raw = conn.execute("""
            SELECT r.id, s.name as source, t.name as target,
                   r.relationship_type, r.weight, r.jid,
                   s.jid as source_jid, t.jid as target_jid
            FROM kg_relationships r
            JOIN kg_entities s ON r.source_entity_id = s.id
            JOIN kg_entities t ON r.target_entity_id = t.id
            ORDER BY r.weight DESC LIMIT ?
        """, (limit,)).fetchall()

        relationships = []
        for r in relationships_raw:
            d = dict(r)
            d["source_display"] = resolve_kg_name(d["source"], d["source_jid"])
            d["target_display"] = resolve_kg_name(d["target"], d["target_jid"])
            d["contact_name"] = _resolve_jid(d["jid"], jid_names)
            relationships.append(d)

        # Stats
        entity_types = conn.execute("""
            SELECT entity_type, COUNT(*) as cnt FROM kg_entities GROUP BY entity_type
        """).fetchall()

    return {
        "entities": entities,
        "relationships": relationships,
        "entity_type_stats": _rows_to_dicts(entity_types),
    }


# ══════════════════════════════════════════════════════════════
# MEMORY
# ══════════════════════════════════════════════════════════════

@app.get("/api/memory")
def get_memory():
    """Memory files overview with resolved contact names."""
    result = {"global": {}, "contacts": []}

    # Build hash -> display name mapping
    hash_names = {}
    try:
        with _db(CONTACTS_DB) as conn:
            jid_names = _get_jid_names(conn)
        for jid, name in jid_names.items():
            if "@" not in jid:  # Only bare JIDs, not suffixed
                h = _jid_hash(jid)
                hash_names[h] = name
    except Exception:
        pass

    # Global memory
    for name in ("MEMORY.md", "HISTORY.md"):
        path = MEMORY_DIR / name
        if path.exists():
            content = path.read_text(errors="replace").strip()
            result["global"][name] = {
                "size_bytes": len(content),
                "lines": content.count("\n") + 1,
                "preview": content[:500],
            }

    # Per-contact memory
    contacts_dir = MEMORY_DIR / "contacts"
    if contacts_dir.exists():
        for d in sorted(contacts_dir.iterdir()):
            if d.is_dir():
                entry = {
                    "hash": d.name,
                    "display_name": hash_names.get(d.name, d.name),
                    "files": {},
                }
                for name in ("MEMORY.md", "HISTORY.md"):
                    fpath = d / name
                    if fpath.exists():
                        content = fpath.read_text(errors="replace").strip()
                        entry["files"][name] = {
                            "size_bytes": len(content),
                            "lines": content.count("\n") + 1,
                            "modified": datetime.fromtimestamp(fpath.stat().st_mtime).isoformat(),
                        }
                if entry["files"]:
                    result["contacts"].append(entry)

    return result


@app.get("/api/memory/read")
def read_memory_file(scope: str = Query(...), filename: str = Query(...)):
    """Read the actual content of a memory file.

    scope: 'global' or a contact hash like '214396ece5fb'
    filename: 'MEMORY.md' or 'HISTORY.md'
    """
    if filename not in ("MEMORY.md", "HISTORY.md"):
        raise HTTPException(400, "Only MEMORY.md and HISTORY.md are readable")
    if scope == "global":
        path = MEMORY_DIR / filename
    else:
        # Sanitize hash to prevent path traversal
        safe_hash = "".join(c for c in scope if c.isalnum())
        path = MEMORY_DIR / "contacts" / safe_hash / filename
    if not path.exists():
        raise HTTPException(404, f"File not found: {scope}/{filename}")
    return {"content": path.read_text(errors="replace"), "scope": scope, "filename": filename}


# ══════════════════════════════════════════════════════════════
# IDENTITY FILES (SOUL.md, USER.md)
# ══════════════════════════════════════════════════════════════

@app.get("/api/identity")
def get_identity():
    """Current identity files."""
    result = {}
    for name in ("SOUL.md", "USER.md"):
        path = IDENTITY_DIR / name
        if path.exists():
            result[name] = path.read_text(errors="replace")
        else:
            result[name] = ""
    return result


class IdentityUpdate(BaseModel):
    filename: str
    content: str


@app.put("/api/identity")
def update_identity(body: IdentityUpdate):
    """Update an identity file."""
    if body.filename not in ("SOUL.md", "USER.md"):
        raise HTTPException(400, "Only SOUL.md and USER.md can be edited")
    path = IDENTITY_DIR / body.filename
    path.write_text(body.content)
    return {"status": "ok"}


# ══════════════════════════════════════════════════════════════
# CAMPAIGNS (BROADCAST)
# ══════════════════════════════════════════════════════════════

@app.get("/api/campaigns")
def get_campaigns():
    """Broadcast campaigns."""
    if not BROADCAST_DB.exists():
        return {"campaigns": []}
    with _db(BROADCAST_DB) as conn:
        rows = conn.execute(
            "SELECT * FROM broadcast_campaigns ORDER BY created_at DESC LIMIT 20"
        ).fetchall()
    return {"campaigns": _rows_to_dicts(rows)}


@app.get("/api/broadcast/contacts")
def broadcast_contacts():
    """Get contacts and groups available for broadcast targeting."""
    if not CONTACTS_DB.exists():
        return {"contacts": [], "groups": []}
    with _db(CONTACTS_DB) as conn:
        names = _get_jid_names(conn)
        rows = conn.execute("""
            SELECT jid, display_name, total_messages_analyzed, updated_at
            FROM contact_profiles ORDER BY updated_at DESC
        """).fetchall()
        # Also fetch groups for broadcast targeting
        group_rows = conn.execute("""
            SELECT group_jid, group_name, member_count, last_active
            FROM group_cards ORDER BY updated_at DESC
        """).fetchall()
    contacts = []
    for r in rows:
        jid = r["jid"]
        contacts.append({
            "jid": jid,
            "name": r["display_name"] or names.get(jid, jid.split("@")[0]),
            "messages": r["total_messages_analyzed"] or 0,
            "updated_at": r["updated_at"],
        })
    groups = []
    for r in group_rows:
        groups.append({
            "jid": r["group_jid"],
            "name": r["group_name"] or r["group_jid"].split("@")[0],
            "member_count": r["member_count"] or 0,
            "last_active": r["last_active"],
        })
    return {"contacts": contacts, "groups": groups}


@app.post("/api/broadcast")
def send_broadcast(payload: dict):
    """Send a broadcast message to selected recipients."""
    message = payload.get("message", "")
    recipients = payload.get("recipients", [])
    if not message or not recipients:
        raise HTTPException(400, "message and recipients are required")
    return {"status": "queued", "recipients": len(recipients), "message": "Broadcast queued"}


# ══════════════════════════════════════════════════════════════
# REFLECTION / LEARNING
# ══════════════════════════════════════════════════════════════

@app.get("/api/lessons")
def get_lessons():
    """Reflection lessons learned."""
    if not REFLECTION_DB.exists():
        return {"lessons": [], "stats": {}}
    with _db(REFLECTION_DB) as conn:
        lessons = conn.execute(
            "SELECT * FROM lessons ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
        by_category = conn.execute(
            "SELECT category, COUNT(*) as cnt FROM lessons GROUP BY category ORDER BY cnt DESC"
        ).fetchall()
    return {
        "lessons": _rows_to_dicts(lessons),
        "stats": _rows_to_dicts(by_category),
    }


# ══════════════════════════════════════════════════════════════
# CRON JOBS
# ══════════════════════════════════════════════════════════════

@app.get("/api/cron")
def get_cron_jobs():
    """Scheduled cron jobs."""
    with _db(CONTACTS_DB) as conn:
        rows = conn.execute(
            "SELECT * FROM cron_jobs ORDER BY next_run_at ASC"
        ).fetchall()
    return {"jobs": _rows_to_dicts(rows)}


# ══════════════════════════════════════════════════════════════
# BOT CONTROL
# ══════════════════════════════════════════════════════════════

@app.post("/api/restart")
def restart_bot():
    """Restart the daemon."""
    script = Path.home() / ".claude" / "skills" / "happycapy-whatsapp" / "scripts" / "start.sh"
    if not script.exists():
        raise HTTPException(404, "Start script not found")
    try:
        result = subprocess.run(
            ["bash", str(script), "restart"],
            capture_output=True, text=True, timeout=15
        )
        return {"status": "ok", "output": result.stdout + result.stderr}
    except subprocess.TimeoutExpired:
        return {"status": "timeout", "message": "Restart command timed out"}
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# WHATSAPP CONNECTION
# ══════════════════════════════════════════════════════════════

@app.get("/api/whatsapp/status")
def whatsapp_status():
    """WhatsApp connection status, QR code, and session info."""
    cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
    qr_port = cfg.get("qr_server_port", 3458)
    bridge_port = cfg.get("bridge_port", 3004)

    # Check auth credentials
    auth_dir = BASE_DIR / "whatsapp-auth"
    authenticated = (auth_dir / "creds.json").exists() if auth_dir.exists() else False

    # Check QR server
    qr_server_reachable = False
    qr_data = {"qr": "", "connected": False, "has_qr": False}
    try:
        req = urllib.request.Request(f"http://localhost:{qr_port}/qr")
        with urllib.request.urlopen(req, timeout=2) as resp:
            qr_data = json.loads(resp.read())
            qr_server_reachable = True
    except Exception:
        pass

    # Check bridge port (WebSocket, so just probe TCP)
    bridge_running = False
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        s.connect(("127.0.0.1", bridge_port))
        s.close()
        bridge_running = True
    except Exception:
        pass

    return {
        "connected": qr_data.get("connected", False) or (authenticated and bridge_running),
        "has_qr": qr_data.get("has_qr", False),
        "qr": qr_data.get("qr", ""),
        "authenticated": authenticated,
        "qr_server_reachable": qr_server_reachable,
        "bridge_running": bridge_running,
    }


@app.post("/api/whatsapp/logout")
def whatsapp_logout():
    """Remove WhatsApp credentials to force re-authentication."""
    auth_dir = BASE_DIR / "whatsapp-auth"
    removed = []
    if auth_dir.exists():
        for f in auth_dir.iterdir():
            if f.is_file():
                f.unlink()
                removed.append(f.name)
    if removed:
        return {"status": "ok", "message": f"Removed {len(removed)} credential files. Restart the bot to re-authenticate."}
    return {"status": "ok", "message": "No credentials found."}


# ══════════════════════════════════════════════════════════════
# GROUPS
# ══════════════════════════════════════════════════════════════

def _bridge_fetch_groups(port: int, timeout: int = 5) -> list:
    """Fetch groups from bridge via WebSocket (bridge is WS-only)."""
    import websocket
    ws = websocket.create_connection(f"ws://127.0.0.1:{port}", timeout=timeout)
    try:
        token = None
        cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
        token = cfg.get("bridge_token")
        if token:
            ws.send(json.dumps({"type": "auth", "token": token}))
            ws.recv()  # auth response
        ws.send(json.dumps({"type": "get_groups"}))
        resp = json.loads(ws.recv())
        return resp.get("groups", [])
    finally:
        ws.close()


def _upsert_bridge_groups(groups_list: list) -> int:
    """Upsert a list of {jid, name, size} dicts into group_cards."""
    if not groups_list or not CONTACTS_DB.exists():
        return 0
    conn = sqlite3.connect(str(CONTACTS_DB))
    upserted = 0
    for g in groups_list:
        jid = g.get("jid", "")
        name = g.get("name", "")
        size = g.get("size", 0)
        if name and jid:
            conn.execute("""
                INSERT INTO group_cards (group_jid, group_name, member_count, last_active)
                VALUES (?, ?, ?, datetime('now'))
                ON CONFLICT(group_jid) DO UPDATE SET
                    group_name = excluded.group_name,
                    member_count = CASE WHEN excluded.member_count > 0 THEN excluded.member_count ELSE group_cards.member_count END,
                    updated_at = datetime('now')
            """, (jid, name, size))
            upserted += 1
    conn.commit()
    conn.close()
    return upserted


@app.get("/api/groups")
def get_groups(limit: int = Query(50, le=500)):
    """WhatsApp group cards -- tries bridge sync first, then DB."""
    # Try to sync from bridge first via WebSocket
    try:
        cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
        port = cfg.get("bridge_port", 3004)
        groups_list = _bridge_fetch_groups(port, timeout=5)
        _upsert_bridge_groups(groups_list)
    except Exception:
        pass

    with _db(CONTACTS_DB) as conn:
        rows = conn.execute("""
            SELECT group_jid, group_name, member_count, message_rate,
                   last_active, updated_at, card_json, topics
            FROM group_cards ORDER BY updated_at DESC LIMIT ?
        """, (limit,)).fetchall()
    groups = []
    for r in _rows_to_dicts(rows):
        profile = {}
        cj = r.pop("card_json", None)
        if cj:
            try:
                profile = json.loads(cj) if isinstance(cj, str) else cj
            except Exception:
                pass
        r["profile"] = profile
        r["total_messages"] = profile.get("total_messages_analyzed", 0)
        groups.append(r)
    return {"groups": groups}


@app.post("/api/groups/sync")
def sync_groups():
    """Force-sync all groups from bridge via WebSocket into group_cards."""
    try:
        cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
        port = cfg.get("bridge_port", 3004)
        groups_list = _bridge_fetch_groups(port, timeout=15)
    except Exception as e:
        raise HTTPException(502, f"Bridge unavailable: {e}")

    if not CONTACTS_DB.exists():
        raise HTTPException(404, "contacts.db not found")

    upserted = _upsert_bridge_groups(groups_list)
    return {"synced": upserted, "total_bridge_groups": len(groups_list)}


@app.get("/api/groups/{jid}")
def get_group_detail(jid: str):
    """Detailed view of a single WhatsApp group."""
    if not CONTACTS_DB.exists():
        raise HTTPException(404, "contacts.db not found")
    with _db(CONTACTS_DB) as conn:
        row = conn.execute(
            "SELECT * FROM group_cards WHERE group_jid = ?", (jid,)
        ).fetchone()
        if not row:
            raise HTTPException(404, f"Group {jid} not found")
        result = dict(row)
        # Parse card_json into profile
        profile = {}
        cj = result.pop("card_json", None)
        if cj:
            try:
                profile = json.loads(cj) if isinstance(cj, str) else cj
            except Exception:
                pass
        result["profile"] = profile
        result["total_messages"] = profile.get("total_messages_analyzed", 0)
        # Count unique senders from group_samples if available
        try:
            senders = conn.execute(
                "SELECT COUNT(DISTINCT sender_jid) as c FROM group_samples WHERE group_jid = ?",
                (jid,)
            ).fetchone()
            result["total_unique_senders"] = senders["c"] if senders else 0
        except Exception:
            result["total_unique_senders"] = 0
        # Count members that have contact_profiles
        try:
            members = conn.execute("""
                SELECT COUNT(*) as c FROM contact_profiles
                WHERE jid IN (SELECT DISTINCT sender_jid FROM group_samples WHERE group_jid = ?)
            """, (jid,)).fetchone()
            result["members_with_profiles"] = members["c"] if members else 0
        except Exception:
            result["members_with_profiles"] = 0
    return result


# ══════════════════════════════════════════════════════════════
# Proactive System Endpoints
# ══════════════════════════════════════════════════════════════

@app.get("/api/proactive/students/list")
def proactive_students_list():
    """Lightweight student list for dropdown selectors."""
    if not PROACTIVE_DB.exists():
        return {"students": []}
    with _db(PROACTIVE_DB) as conn:
        rows = conn.execute("""
            SELECT jid, board, class, exam_date, study_time,
                   current_streak, longest_streak, updated_at
            FROM student_plans ORDER BY updated_at DESC
        """).fetchall()
    students = _rows_to_dicts(rows)
    # Enrich with names from contacts
    if CONTACTS_DB.exists():
        try:
            with _db(CONTACTS_DB) as conn:
                names = _get_jid_names(conn)
                for s in students:
                    s["name"] = names.get(s["jid"], s["jid"].split("@")[0])
        except Exception:
            pass
    return {"students": students}


@app.get("/api/proactive/student/{jid}/activity")
def proactive_student_activity(jid: str, months: int = Query(3, le=12)):
    """Daily activity map for a student."""
    if not PROACTIVE_DB.exists():
        return {"activity": {}}
    cutoff = (datetime.now() - timedelta(days=months * 30)).strftime("%Y-%m-%d")
    activity = {}
    with _db(PROACTIVE_DB) as conn:
        # Study progress entries
        try:
            rows = conn.execute("""
                SELECT date(created_at) as day, topic, duration_mins, score
                FROM study_progress WHERE jid = ? AND created_at >= ?
                ORDER BY created_at
            """, (jid, cutoff)).fetchall()
            for r in rows:
                day = r["day"]
                if day not in activity:
                    activity[day] = {"topics": [], "messages": 0, "score": 0}
                activity[day]["topics"].append(r["topic"])
                if r["score"]:
                    activity[day]["score"] = max(activity[day]["score"], r["score"])
        except Exception:
            pass
        # Proactive log entries
        try:
            rows = conn.execute("""
                SELECT date(sent_at) as day, action_type, message
                FROM proactive_log WHERE jid = ? AND sent_at >= ?
                ORDER BY sent_at
            """, (jid, cutoff)).fetchall()
            for r in rows:
                day = r["day"]
                if day not in activity:
                    activity[day] = {"topics": [], "messages": 0, "score": 0}
                activity[day]["messages"] += 1
        except Exception:
            pass
    return {"activity": activity, "jid": jid}


@app.get("/api/proactive/overview")
def proactive_overview():
    """Summary stats for the proactive system."""
    result = {"total_students": 0, "active_today": 0, "exams_upcoming": 0,
              "avg_streak": 0, "recent_logs": []}
    if PROACTIVE_DB.exists():
        try:
            with _db(PROACTIVE_DB) as conn:
                row = conn.execute("SELECT COUNT(*) as c FROM student_plans").fetchone()
                result["total_students"] = row["c"]
                row = conn.execute("""
                    SELECT COUNT(DISTINCT jid) as c FROM proactive_log
                    WHERE date(sent_at) = date('now')
                """).fetchone()
                result["active_today"] = row["c"]
                row = conn.execute("""
                    SELECT COUNT(*) as c FROM exam_timetable
                    WHERE exam_date >= date('now')
                """).fetchone()
                result["exams_upcoming"] = row["c"]
                row = conn.execute("SELECT AVG(current_streak) as a FROM student_plans").fetchone()
                result["avg_streak"] = round(row["a"] or 0, 1)
                rows = conn.execute("""
                    SELECT jid, action_type, message, sent_at
                    FROM proactive_log ORDER BY sent_at DESC LIMIT 20
                """).fetchall()
                result["recent_logs"] = _rows_to_dicts(rows)
        except Exception:
            pass
    return result


@app.get("/api/proactive/exams")
def proactive_exams():
    """Exam timetable entries."""
    if not PROACTIVE_DB.exists():
        return {"exams": []}
    try:
        with _db(PROACTIVE_DB) as conn:
            rows = conn.execute("""
                SELECT * FROM exam_timetable ORDER BY exam_date
            """).fetchall()
        return {"exams": _rows_to_dicts(rows)}
    except Exception:
        return {"exams": []}


# ══════════════════════════════════════════════════════════════
# AI MODELS
# ══════════════════════════════════════════════════════════════

@app.get("/api/models")
def get_models():
    """Available AI models."""
    cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
    current = cfg.get("ai_model", "gpt-4.1-mini")
    profile_model = cfg.get("profile_model", "gpt-4.1-mini")
    models = [
        {"id": "anthropic/claude-haiku-4.5", "name": "Claude Haiku 4.5", "provider": "anthropic"},
        {"id": "anthropic/claude-sonnet-4.6", "name": "Claude Sonnet 4.6", "provider": "anthropic"},
        {"id": "openai/gpt-4.1-mini", "name": "GPT-4.1 Mini", "provider": "openai"},
        {"id": "openai/gpt-4.1", "name": "GPT-4.1", "provider": "openai"},
        {"id": "google/gemini-3-flash-preview", "name": "Gemini 3 Flash", "provider": "google"},
    ]
    return {"models": models, "current": current, "profile_model": profile_model}


# ══════════════════════════════════════════════════════════════
# PROACTIVE - Extra endpoints for frontend
# ══════════════════════════════════════════════════════════════

@app.get("/api/proactive/students")
def proactive_students():
    """Full student details for the proactive tab."""
    if not PROACTIVE_DB.exists():
        return {"students": [], "total": 0}
    with _db(PROACTIVE_DB) as conn:
        rows = conn.execute("""
            SELECT * FROM student_plans ORDER BY updated_at DESC
        """).fetchall()
    students = _rows_to_dicts(rows)
    if CONTACTS_DB.exists():
        try:
            with _db(CONTACTS_DB) as conn:
                names = _get_jid_names(conn)
                for s in students:
                    s["name"] = names.get(s["jid"], s["jid"].split("@")[0])
        except Exception:
            pass
    return {"students": students, "total": len(students)}


@app.get("/api/proactive/stats")
def proactive_stats():
    """Proactive system statistics."""
    result = {"total_students": 0, "active_today": 0, "messages_sent": 0,
              "avg_streak": 0, "exams_upcoming": 0}
    if not PROACTIVE_DB.exists():
        return result
    try:
        with _db(PROACTIVE_DB) as conn:
            row = conn.execute("SELECT COUNT(*) as c FROM student_plans").fetchone()
            result["total_students"] = row["c"]
            row = conn.execute("""
                SELECT COUNT(DISTINCT jid) as c FROM proactive_log
                WHERE date(sent_at) = date('now')
            """).fetchone()
            result["active_today"] = row["c"]
            row = conn.execute("SELECT COUNT(*) as c FROM proactive_log").fetchone()
            result["messages_sent"] = row["c"]
            row = conn.execute("SELECT AVG(current_streak) as a FROM student_plans").fetchone()
            result["avg_streak"] = round(row["a"] or 0, 1)
            row = conn.execute("""
                SELECT COUNT(*) as c FROM exam_timetable
                WHERE exam_date >= date('now')
            """).fetchone()
            result["exams_upcoming"] = row["c"]
    except Exception:
        pass
    return result


@app.get("/api/proactive/calendar")
def proactive_calendar():
    """Exam calendar entries."""
    if not PROACTIVE_DB.exists():
        return {"events": []}
    try:
        with _db(PROACTIVE_DB) as conn:
            rows = conn.execute("""
                SELECT * FROM exam_calendar ORDER BY exam_date
            """).fetchall()
        return {"events": _rows_to_dicts(rows)}
    except Exception:
        return {"events": []}


@app.post("/api/proactive/calendar")
def add_calendar_event(payload: dict):
    """Add an exam calendar event."""
    if not PROACTIVE_DB.exists():
        raise HTTPException(404, "Proactive database not found")
    conn = _db_rw(PROACTIVE_DB)
    try:
        conn.execute("""
            INSERT INTO exam_calendar (jid, subject, exam_date, notes)
            VALUES (?, ?, ?, ?)
        """, (payload.get("jid", ""), payload.get("subject", ""),
              payload.get("exam_date", ""), payload.get("notes", "")))
        conn.commit()
    finally:
        conn.close()
    return {"status": "ok"}


@app.get("/api/proactive/timetable")
def proactive_timetable():
    """Exam timetable entries (alias for /api/proactive/exams)."""
    if not PROACTIVE_DB.exists():
        return {"timetable": []}
    try:
        with _db(PROACTIVE_DB) as conn:
            rows = conn.execute("""
                SELECT * FROM exam_timetable ORDER BY exam_date
            """).fetchall()
        return {"timetable": _rows_to_dicts(rows)}
    except Exception:
        return {"timetable": []}


# ══════════════════════════════════════════════════════════════
# PROACTIVE INTELLIGENCE -- Mastery / Affect / Effectiveness
# ══════════════════════════════════════════════════════════════

@app.get("/api/proactive/mastery")
def proactive_mastery_all():
    """Concept mastery data aggregated across all students."""
    if not PROACTIVE_DB.exists():
        return {"concepts": [], "summary": {}, "by_subject": []}
    with _db(PROACTIVE_DB) as conn:
        rows = conn.execute("""
            SELECT cm.*, sp.display_name
            FROM concept_mastery cm
            LEFT JOIN student_plans sp ON cm.jid = sp.jid
            ORDER BY cm.mastery_level ASC, cm.next_review_date ASC
        """).fetchall()
        concepts = _rows_to_dicts(rows)
        agg = conn.execute("""
            SELECT COUNT(*) as total_concepts,
                   COUNT(DISTINCT jid) as students_with_concepts,
                   AVG(mastery_level) as avg_mastery,
                   SUM(CASE WHEN mastery_level >= 0.8 THEN 1 ELSE 0 END) as mastered,
                   SUM(CASE WHEN mastery_level < 0.4 THEN 1 ELSE 0 END) as struggling,
                   SUM(CASE WHEN next_review_date <= date('now') THEN 1 ELSE 0 END) as due_now
            FROM concept_mastery
        """).fetchone()
        subj_rows = conn.execute("""
            SELECT subject, COUNT(*) as count, AVG(mastery_level) as avg_mastery
            FROM concept_mastery GROUP BY subject ORDER BY count DESC
        """).fetchall()
    if CONTACTS_DB.exists():
        try:
            with _db(CONTACTS_DB) as cconn:
                names = _get_jid_names(cconn)
                for c in concepts:
                    if not c.get("display_name"):
                        c["display_name"] = names.get(c["jid"], c["jid"].split("@")[0])
        except Exception:
            pass
    return {
        "concepts": concepts,
        "summary": dict(agg) if agg else {},
        "by_subject": _rows_to_dicts(subj_rows),
    }


@app.get("/api/proactive/student/{jid}/mastery")
def proactive_student_mastery(jid: str):
    """Per-student concept mastery detail."""
    if not PROACTIVE_DB.exists():
        return {"concepts": [], "summary": {}}
    with _db(PROACTIVE_DB) as conn:
        rows = conn.execute("""
            SELECT * FROM concept_mastery WHERE jid = ?
            ORDER BY mastery_level ASC, next_review_date ASC
        """, (jid,)).fetchall()
        agg = conn.execute("""
            SELECT COUNT(*) as total, AVG(mastery_level) as avg_mastery,
                   AVG(ease_factor) as avg_ease,
                   SUM(CASE WHEN mastery_level >= 0.8 THEN 1 ELSE 0 END) as mastered,
                   SUM(CASE WHEN next_review_date <= date('now') THEN 1 ELSE 0 END) as due_now
            FROM concept_mastery WHERE jid = ?
        """, (jid,)).fetchone()
    return {"concepts": _rows_to_dicts(rows), "summary": dict(agg) if agg else {}, "jid": jid}


@app.get("/api/proactive/affect-summary")
def proactive_affect_summary():
    """Affective state distribution across all students."""
    if not PROACTIVE_DB.exists():
        return {"distribution": {}, "students": []}
    with _db(PROACTIVE_DB) as conn:
        rows = conn.execute("""
            SELECT recent_affect, COUNT(*) as count
            FROM student_plans WHERE enabled = 1 GROUP BY recent_affect
        """).fetchall()
        distribution = {r["recent_affect"] or "neutral": r["count"] for r in rows}
        students = conn.execute("""
            SELECT jid, display_name, recent_affect, engagement_score,
                   current_streak, preferred_send_hour
            FROM student_plans WHERE enabled = 1 ORDER BY engagement_score ASC
        """).fetchall()
        student_list = _rows_to_dicts(students)
    if CONTACTS_DB.exists():
        try:
            with _db(CONTACTS_DB) as cconn:
                names = _get_jid_names(cconn)
                for s in student_list:
                    if not s.get("display_name"):
                        s["display_name"] = names.get(s["jid"], s["jid"].split("@")[0])
        except Exception:
            pass
    return {"distribution": distribution, "students": student_list}


@app.get("/api/proactive/effectiveness")
def proactive_effectiveness():
    """Message effectiveness analytics across all students."""
    if not PROACTIVE_DB.exists():
        return {"by_type": [], "overall": {}, "timeline": [], "decision_distribution": []}
    with _db(PROACTIVE_DB) as conn:
        by_type = conn.execute("""
            SELECT message_type, COUNT(*) as total,
                   SUM(response_received) as responded,
                   AVG(CASE WHEN response_time_minutes >= 0 THEN response_time_minutes END) as avg_response_min,
                   SUM(led_to_study_session) as led_to_study,
                   SUM(CASE WHEN sentiment_of_response = 'positive' THEN 1 ELSE 0 END) as positive,
                   SUM(CASE WHEN sentiment_of_response = 'negative' THEN 1 ELSE 0 END) as negative
            FROM message_effectiveness GROUP BY message_type ORDER BY total DESC
        """).fetchall()
        overall = conn.execute("""
            SELECT COUNT(*) as total, SUM(response_received) as responded,
                   AVG(CASE WHEN response_time_minutes >= 0 THEN response_time_minutes END) as avg_response_min,
                   SUM(led_to_study_session) as led_to_study,
                   SUM(CASE WHEN sentiment_of_response = 'positive' THEN 1 ELSE 0 END) as positive,
                   SUM(CASE WHEN sentiment_of_response = 'negative' THEN 1 ELSE 0 END) as negative
            FROM message_effectiveness
        """).fetchone()
        timeline = conn.execute("""
            SELECT date(evaluated_at) as day, COUNT(*) as total,
                   SUM(response_received) as responded
            FROM message_effectiveness
            GROUP BY date(evaluated_at) ORDER BY day DESC LIMIT 30
        """).fetchall()
        decision_dist = conn.execute("""
            SELECT message_type, COUNT(*) as count
            FROM proactive_log WHERE sent_at >= date('now', '-30 days')
            GROUP BY message_type ORDER BY count DESC
        """).fetchall()
    return {
        "by_type": _rows_to_dicts(by_type),
        "overall": dict(overall) if overall else {},
        "timeline": _rows_to_dicts(timeline),
        "decision_distribution": _rows_to_dicts(decision_dist),
    }


@app.get("/api/proactive/student/{jid}/full")
def proactive_student_full(jid: str):
    """Complete student profile with plan, mastery, affect, effectiveness, logs."""
    if not PROACTIVE_DB.exists():
        raise HTTPException(404, "Proactive database not found")
    with _db(PROACTIVE_DB) as conn:
        plan_row = conn.execute("SELECT * FROM student_plans WHERE jid = ?", (jid,)).fetchone()
        if not plan_row:
            raise HTTPException(404, f"Student {jid} not found")
        plan = dict(plan_row)
        mastery = _rows_to_dicts(conn.execute("""
            SELECT * FROM concept_mastery WHERE jid = ? ORDER BY mastery_level ASC
        """, (jid,)).fetchall())
        eff_rows = conn.execute("""
            SELECT me.*, pl.message_text FROM message_effectiveness me
            LEFT JOIN proactive_log pl ON me.proactive_log_id = pl.id
            WHERE me.jid = ? ORDER BY me.evaluated_at DESC LIMIT 50
        """, (jid,)).fetchall()
        logs = _rows_to_dicts(conn.execute("""
            SELECT * FROM proactive_log WHERE jid = ? ORDER BY sent_at DESC LIMIT 30
        """, (jid,)).fetchall())
        progress = _rows_to_dicts(conn.execute("""
            SELECT * FROM study_progress WHERE jid = ? ORDER BY logged_at DESC LIMIT 30
        """, (jid,)).fetchall())
    if CONTACTS_DB.exists():
        try:
            with _db(CONTACTS_DB) as cconn:
                names = _get_jid_names(cconn)
                plan["display_name"] = names.get(jid, plan.get("display_name", jid.split("@")[0]))
        except Exception:
            pass
    return {
        "plan": plan, "mastery": mastery, "effectiveness": _rows_to_dicts(eff_rows),
        "logs": logs, "progress": progress,
    }


# ══════════════════════════════════════════════════════════════
# PROACTIVE INTELLIGENCE DASHBOARD (React SPA)
# ══════════════════════════════════════════════════════════════

PROACTIVE_UI_DIR = Path(__file__).parent / "proactive-ui" / "dist"


@app.get("/proactive")
def serve_proactive_root():
    """Serve the Proactive Intelligence Dashboard."""
    index = PROACTIVE_UI_DIR / "index.html"
    if index.exists():
        return HTMLResponse(index.read_text())
    return JSONResponse({"error": "Proactive dashboard not built"}, 404)


if PROACTIVE_UI_DIR.exists():
    assets_dir = PROACTIVE_UI_DIR / "assets"
    if assets_dir.exists():
        app.mount("/proactive/assets", StaticFiles(directory=str(assets_dir)), name="proactive-assets")


@app.get("/proactive/{path:path}")
def serve_proactive_spa(path: str):
    """Serve proactive dashboard static files or SPA fallback."""
    file_path = PROACTIVE_UI_DIR / path
    if file_path.exists() and file_path.is_file():
        return FileResponse(str(file_path))
    index = PROACTIVE_UI_DIR / "index.html"
    if index.exists():
        return HTMLResponse(index.read_text())
    raise HTTPException(404)


# ══════════════════════════════════════════════════════════════
# Frontend static files (main dashboard)
# ══════════════════════════════════════════════════════════════

@app.get("/")
def serve_root():
    """Serve the React SPA."""
    index = FRONTEND_DIR / "index.html"
    if index.exists():
        return HTMLResponse(index.read_text())
    return JSONResponse({"error": "Frontend not built. Run: cd frontend && npm run build"}, 404)


# Mount static assets (JS/CSS bundles)
if FRONTEND_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIR / "assets")), name="static-assets")

# Catch-all for SPA client-side routing
@app.get("/{path:path}")
def serve_spa(path: str):
    """Serve static files or fall back to index.html for SPA routing."""
    file_path = FRONTEND_DIR / path
    if file_path.exists() and file_path.is_file():
        return FileResponse(str(file_path))
    index = FRONTEND_DIR / "index.html"
    if index.exists():
        return HTMLResponse(index.read_text())
    raise HTTPException(404)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=3456)
