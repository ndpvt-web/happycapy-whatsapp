"""Spreadsheet Tracker integration for HappyCapy WhatsApp bot.

Lets the LLM log, read, update, and send Excel spreadsheets.
Use cases: order tracking, expense logging, customer data, inventory management.

Files stored in ~/.happycapy-whatsapp/data/spreadsheets/{name}.xlsx
"""

import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .base import BaseIntegration, IntegrationInfo

# Import ToolResult from the parent tool_executor module
from src.tool_executor import ToolResult


# ── Tool Definitions (OpenAI format) ──

_TOOL_DEFINITIONS: list[dict] = [
    {
        "type": "function",
        "function": {
            "name": "log_to_spreadsheet",
            "description": (
                "Add a new row of data to a named Excel spreadsheet. Creates the spreadsheet "
                "if it doesn't exist. Use for logging orders, expenses, customers, inventory, or any structured data."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "spreadsheet_name": {
                        "type": "string",
                        "description": "Name of the spreadsheet (e.g. 'orders', 'expenses', 'customers'). Letters, numbers, hyphens, underscores only.",
                    },
                    "data": {
                        "type": "object",
                        "description": "Key-value pairs for the row. Keys become column headers. Example: {\"customer\": \"John\", \"items\": \"5 widgets\", \"amount\": 150, \"status\": \"pending\"}",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet/tab name within the spreadsheet. Default: 'Sheet1'.",
                    },
                },
                "required": ["spreadsheet_name", "data"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_spreadsheet",
            "description": (
                "Read recent rows from a spreadsheet. Returns data as a formatted table. "
                "Can optionally filter by a column value."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "spreadsheet_name": {
                        "type": "string",
                        "description": "Name of the spreadsheet to read.",
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Maximum number of rows to return. Default: 10.",
                    },
                    "filter_column": {
                        "type": "string",
                        "description": "Column name to filter by (optional).",
                    },
                    "filter_value": {
                        "type": "string",
                        "description": "Value to match in filter_column (case-insensitive partial match).",
                    },
                },
                "required": ["spreadsheet_name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_spreadsheet_row",
            "description": (
                "Update specific columns in an existing spreadsheet row. "
                "Row numbers match Excel (row 1 = headers, row 2 = first data row)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "spreadsheet_name": {
                        "type": "string",
                        "description": "Name of the spreadsheet.",
                    },
                    "row_number": {
                        "type": "integer",
                        "description": "Row number to update (2 = first data row, since row 1 is headers).",
                    },
                    "updates": {
                        "type": "object",
                        "description": "Column name -> new value pairs. Example: {\"status\": \"shipped\", \"notes\": \"Tracking #12345\"}",
                    },
                },
                "required": ["spreadsheet_name", "row_number", "updates"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "send_spreadsheet",
            "description": (
                "Send a spreadsheet file to a WhatsApp contact as a document attachment."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "spreadsheet_name": {
                        "type": "string",
                        "description": "Name of the spreadsheet to send.",
                    },
                    "phone_number": {
                        "type": "string",
                        "description": "Phone number with country code, digits only (e.g. '14155551234').",
                    },
                },
                "required": ["spreadsheet_name", "phone_number"],
            },
        },
    },
]


# ── Spreadsheet Integration ──

_SAFE_NAME_RE = re.compile(r"^[a-zA-Z0-9_-]+$")
_MAX_ROWS = 10_000


class Integration(BaseIntegration):
    """Spreadsheet tracking integration."""

    DATA_DIR = Path.home() / ".happycapy-whatsapp" / "data" / "spreadsheets"
    MEDIA_DIR = Path.home() / ".happycapy-whatsapp" / "media"

    def __init__(self, config: dict[str, Any], **kwargs: Any):
        self.config = config
        self._channel = kwargs.get("channel")
        self.DATA_DIR.mkdir(parents=True, exist_ok=True)
        self.MEDIA_DIR.mkdir(parents=True, exist_ok=True)

    @classmethod
    def info(cls) -> IntegrationInfo:
        return IntegrationInfo(
            name="spreadsheet",
            display_name="Spreadsheet Tracker",
            description="Log and track structured data in Excel spreadsheets",
        )

    @classmethod
    def tool_definitions(cls) -> list[dict]:
        return _TOOL_DEFINITIONS

    @classmethod
    def system_prompt_addition(cls, config: dict[str, Any]) -> str:
        return (
            "## Spreadsheet Tracking\n"
            "You can log structured data to Excel spreadsheets. Use for orders, expenses, customers, inventory, etc.\n"
            "Tools: log_to_spreadsheet, read_spreadsheet, update_spreadsheet_row, send_spreadsheet\n"
            "- When someone places an order or gives you data to track, use log_to_spreadsheet\n"
            "- When asked about past data, use read_spreadsheet\n"
            "- Use consistent column names across entries (e.g. customer, items, amount, status)\n"
            "- Each spreadsheet auto-tracks timestamps"
        )

    async def execute(self, tool_name: str, arguments: dict[str, Any]) -> ToolResult:
        """Execute a spreadsheet tool."""
        handlers = {
            "log_to_spreadsheet": self._log,
            "read_spreadsheet": self._read,
            "update_spreadsheet_row": self._update,
            "send_spreadsheet": self._send,
        }
        handler = handlers.get(tool_name)
        if not handler:
            return ToolResult(False, tool_name, f"Unknown spreadsheet tool: {tool_name}")
        try:
            return await handler(arguments)
        except Exception as e:
            return ToolResult(False, tool_name, f"Spreadsheet error: {type(e).__name__}: {e}")

    # ── Helpers ──

    def _get_path(self, name: str) -> Path | None:
        """Get sanitized spreadsheet file path. Returns None if name is invalid."""
        name = name.strip()
        if not name or not _SAFE_NAME_RE.match(name):
            return None
        return self.DATA_DIR / f"{name}.xlsx"

    def _get_or_create_wb(self, path: Path, sheet_name: str = "Sheet1") -> tuple[Workbook, Any]:
        """Load existing workbook or create a new one. Returns (wb, sheet)."""
        if path.exists():
            wb = load_workbook(str(path))
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
        return wb, ws

    def _get_headers(self, ws: Any) -> list[str]:
        """Get column headers from row 1."""
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                headers.append(str(val))
            else:
                break
        return headers

    # ── Tool Handlers ──

    async def _log(self, args: dict[str, Any]) -> ToolResult:
        """Add a new row to a spreadsheet."""
        name = args.get("spreadsheet_name", "")
        path = self._get_path(name)
        if not path:
            return ToolResult(False, "log_to_spreadsheet", "Invalid spreadsheet name. Use only letters, numbers, hyphens, underscores.")

        data = args.get("data", {})
        if not data or not isinstance(data, dict):
            return ToolResult(False, "log_to_spreadsheet", "No data provided. Pass key-value pairs like {\"customer\": \"John\", \"amount\": 150}")

        sheet_name = args.get("sheet_name", "Sheet1") or "Sheet1"
        wb, ws = self._get_or_create_wb(path, sheet_name)

        # Check row limit
        if ws.max_row > _MAX_ROWS:
            wb.close()
            return ToolResult(False, "log_to_spreadsheet", f"Spreadsheet has {ws.max_row} rows (max {_MAX_ROWS}). Create a new spreadsheet.")

        # Build headers: timestamp first, then existing headers, then new keys
        existing_headers = self._get_headers(ws)
        if not existing_headers:
            # New sheet: create headers from data keys with timestamp first
            headers = ["timestamp"] + [k for k in data.keys() if k != "timestamp"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            existing_headers = headers

        # Ensure all data keys have a column (add new columns if needed)
        for key in data.keys():
            if key not in existing_headers and key != "timestamp":
                new_col = len(existing_headers) + 1
                ws.cell(row=1, column=new_col, value=key)
                existing_headers.append(key)

        # Add row
        row_num = ws.max_row + 1
        # Always set timestamp
        if "timestamp" in existing_headers:
            col_idx = existing_headers.index("timestamp") + 1
            ws.cell(row=row_num, column=col_idx, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        for key, value in data.items():
            if key in existing_headers:
                col_idx = existing_headers.index(key) + 1
                ws.cell(row=row_num, column=col_idx, value=value)

        wb.save(str(path))
        wb.close()

        return ToolResult(
            True, "log_to_spreadsheet",
            f"Row {row_num - 1} added to '{name}' spreadsheet ({len(data)} columns). Total rows: {row_num - 1}.",
        )

    async def _read(self, args: dict[str, Any]) -> ToolResult:
        """Read rows from a spreadsheet."""
        name = args.get("spreadsheet_name", "")
        path = self._get_path(name)
        if not path:
            return ToolResult(False, "read_spreadsheet", "Invalid spreadsheet name.")
        if not path.exists():
            return ToolResult(False, "read_spreadsheet", f"Spreadsheet '{name}' does not exist yet. Use log_to_spreadsheet to create it.")

        limit = min(args.get("limit", 10), 50)  # Cap at 50
        filter_col = args.get("filter_column", "")
        filter_val = args.get("filter_value", "")

        wb = load_workbook(str(path), read_only=True)
        ws = wb.active

        headers = self._get_headers(ws)
        if not headers:
            wb.close()
            return ToolResult(False, "read_spreadsheet", f"Spreadsheet '{name}' is empty.")

        # Read all data rows
        rows: list[dict[str, Any]] = []
        for row_idx in range(2, ws.max_row + 1):
            row_data: dict[str, Any] = {"_row": row_idx}
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = val if val is not None else ""
            rows.append(row_data)

        wb.close()

        # Apply filter
        if filter_col and filter_val and filter_col in headers:
            filter_val_lower = filter_val.lower()
            rows = [r for r in rows if filter_val_lower in str(r.get(filter_col, "")).lower()]

        # Take most recent rows (last N)
        total = len(rows)
        rows = rows[-limit:]

        if not rows:
            msg = f"No rows found in '{name}'"
            if filter_col:
                msg += f" matching {filter_col}='{filter_val}'"
            return ToolResult(True, "read_spreadsheet", msg)

        # Format as text table
        display_headers = ["Row#"] + headers
        lines = [" | ".join(display_headers)]
        lines.append("-" * len(lines[0]))
        for row in rows:
            vals = [str(row["_row"])] + [str(row.get(h, "")) for h in headers]
            lines.append(" | ".join(vals))

        result_text = f"Spreadsheet '{name}' ({total} total rows, showing {len(rows)}):\n\n" + "\n".join(lines)
        return ToolResult(True, "read_spreadsheet", result_text)

    async def _update(self, args: dict[str, Any]) -> ToolResult:
        """Update a specific row in a spreadsheet."""
        name = args.get("spreadsheet_name", "")
        path = self._get_path(name)
        if not path:
            return ToolResult(False, "update_spreadsheet_row", "Invalid spreadsheet name.")
        if not path.exists():
            return ToolResult(False, "update_spreadsheet_row", f"Spreadsheet '{name}' does not exist.")

        row_num = args.get("row_number", 0)
        updates = args.get("updates", {})
        if not updates or not isinstance(updates, dict):
            return ToolResult(False, "update_spreadsheet_row", "No updates provided.")

        wb = load_workbook(str(path))
        ws = wb.active

        headers = self._get_headers(ws)
        if not headers:
            wb.close()
            return ToolResult(False, "update_spreadsheet_row", "Spreadsheet is empty.")

        if row_num < 2 or row_num > ws.max_row:
            wb.close()
            return ToolResult(False, "update_spreadsheet_row", f"Row {row_num} out of range (2-{ws.max_row}).")

        updated_cols = []
        for key, value in updates.items():
            if key in headers:
                col_idx = headers.index(key) + 1
                ws.cell(row=row_num, column=col_idx, value=value)
                updated_cols.append(key)

        if not updated_cols:
            wb.close()
            return ToolResult(False, "update_spreadsheet_row", f"None of the columns {list(updates.keys())} exist in this spreadsheet. Available: {headers}")

        wb.save(str(path))
        wb.close()

        return ToolResult(
            True, "update_spreadsheet_row",
            f"Row {row_num} in '{name}' updated: {', '.join(f'{c}={updates[c]}' for c in updated_cols)}",
        )

    async def _send(self, args: dict[str, Any]) -> ToolResult:
        """Send a spreadsheet file via WhatsApp."""
        if not self._channel:
            return ToolResult(False, "send_spreadsheet", "WhatsApp channel not available.")

        name = args.get("spreadsheet_name", "")
        path = self._get_path(name)
        if not path:
            return ToolResult(False, "send_spreadsheet", "Invalid spreadsheet name.")
        if not path.exists():
            return ToolResult(False, "send_spreadsheet", f"Spreadsheet '{name}' does not exist.")

        phone = "".join(c for c in args.get("phone_number", "") if c.isdigit())
        if not phone or len(phone) < 7:
            return ToolResult(False, "send_spreadsheet", "Invalid phone number.")

        # Copy to media dir (send_media validates paths are within MEDIA_DIR)
        media_path = self.MEDIA_DIR / f"{name}_{int(time.time())}.xlsx"
        shutil.copy2(str(path), str(media_path))

        chat_jid = f"{phone}@s.whatsapp.net"
        try:
            await self._channel.send_media(chat_jid, str(media_path))
        except Exception as e:
            media_path.unlink(missing_ok=True)
            return ToolResult(False, "send_spreadsheet", f"Failed to send: {type(e).__name__}")

        return ToolResult(True, "send_spreadsheet", f"Spreadsheet '{name}' sent to {phone}.", media_path=str(media_path))
